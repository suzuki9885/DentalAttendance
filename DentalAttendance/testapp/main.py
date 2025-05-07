from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import case
import os
from datetime import datetime, timedelta, timezone
from testapp import app, db
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO, StringIO
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase import pdfmetrics
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from testapp.config import ADMIN_ID, ADMIN_PASSWORD

from testapp.models import AttendanceRecord, User

#先頭URLにリクエストが来るとemployee_login.htmlを表示
@app.route('/')
def index():
    return render_template('testapp/employee_login.html')

#管理者ログインボタンの処理
@app.route('/adm_login', methods=['GET', 'POST'])
def adm_login():
    #管理者ログインボタンを押したら
    if request.method == 'POST':
        #IDとパスワードを取得
        adm_id = request.form.get('adm_id')
        adm_password = request.form.get('adm_password')

        # IDとパスワードが一致した場合はmanagement.htmlを表示
        if adm_id == ADMIN_ID and adm_password == ADMIN_PASSWORD:
            session['adm_logged_in'] = True
            return redirect(url_for('management'))
        # 一致しなかった場合は再度adm_login.htmlを表示
        else:
            return render_template('testapp/adm_login.html')
    #管理者ログインボタンを押していない場合はadm_login.htmlを表示
    return render_template('testapp/adm_login.html')

#従業員ログインボタンの処理
@app.route('/employee_login', methods=['GET', 'POST'])
def employee_login():
    if request.method == 'POST':
        employee_id = request.form.get('employee_id')
        password = request.form.get('password')
        
        user = User.query.filter_by(employee_id=employee_id).first()
        
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id  # ユーザーIDをセッションに保存
            session['employee_id'] = employee_id  # 従業員IDをセッションに保存
            session['employment_type'] = user.employment_type  # 雇用形態をセッションに保存
            session['name'] = user.name  # 名前をセッションに保存
            if user.employment_type == 'PT':
                return redirect(url_for('attendance_management_PT', employee_id=employee_id))
            else:
                return redirect(url_for('attendance_management_FT', employee_id=employee_id))
        
        return render_template('testapp/employee_login.html')
    
    return render_template('testapp/employee_login.html')

# 正社員用の勤怠登録画面
@app.route('/attendance_management_FT')
def attendance_management_FT():
    if 'user_id' not in session:
        return redirect(url_for('employee_login'))
    return render_template('testapp/attendance_management_FT.html')

# パート用の勤怠登録画面
@app.route('/attendance_management_PT', methods=['GET', 'POST'])
def attendance_management_PT():
    if 'user_id' not in session:
        return redirect(url_for('employee_login'))
    return render_template('testapp/attendance_management_PT.html')

# 管理者用画面
@app.route('/management', methods=['GET', 'POST'])
def management():
    if 'adm_logged_in' not in session:
        return redirect(url_for('adm_login'))
    return render_template('testapp/management.html')

# アカウント追加画面
@app.route('/add_account', methods=['GET', 'POST'])
def add_account():
    if 'adm_logged_in' not in session:
        return redirect(url_for('adm_login'))
    if request.method == 'POST':
        try:
            # フォームデータの取得
            name = request.form.get('name')
            employment_type = request.form.get('employment_type')
            employee_id = request.form.get('employee_id')
            password = request.form.get('password')

            # 入力値の検証
            if not all([name, employment_type, employee_id, password]):
                flash('すべての項目を入力してください。', 'error')
                return redirect(url_for('add_account'))

            # 従業員IDの重複チェック
            existing_user = User.query.filter_by(employee_id=employee_id).first()
            if existing_user:
                flash('既に存在している従業員IDです', 'error')
                return redirect(url_for('add_account'))

            # パスワードのハッシュ化
            hashed_password = generate_password_hash(password)

            # 新規ユーザーの作成
            new_user = User(
                name=name,
                employment_type=employment_type,
                employee_id=employee_id,
                password=hashed_password
            )

            # データベースに保存
            db.session.add(new_user)
            db.session.commit()

            # 成功メッセージをフラッシュ
            flash('アカウントが正常に作成されました。', 'success')
            # 成功時はアカウント管理画面にリダイレクト
            return redirect(url_for('management_account'))

        except Exception as e:
            # エラーが発生した場合はロールバック
            db.session.rollback()
            # エラーメッセージをフラッシュ（エラー内容を含める）
            flash(f'アカウントの作成に失敗しました: {str(e)}', 'error')
            return redirect(url_for('add_account'))

    # GETリクエストの場合はアカウント追加ページを表示
    return render_template('testapp/add_account.html')

# アカウント管理画面
@app.route('/management_account')
def management_account():
    if 'adm_logged_in' not in session:
        return redirect(url_for('adm_login'))
    
    # 雇用形態（FT/PT）と従業員IDでソート
    users = User.query.order_by(
        case(
            (User.employment_type == 'FT', 0),
            (User.employment_type == 'PT', 1),
        ),
        User.employee_id
    ).all()
    
    # 従業員IDの数字部分でソート
    users = sorted(users, key=lambda x: (x.employment_type, int(x.employee_id[2:]) if x.employee_id[2:].isdigit() else 0))
    
    return render_template('testapp/management_account.html', users=users)

# 勤怠履歴画面
@app.route('/attendance_history')
@app.route('/attendance_history/<int:year>/<int:month>')
def attendance_history(year=None, month=None):
    # パラメータが指定されていない場合は現在の日付を使用
    if year is None or month is None:
        now = datetime.now()
        # 21日以降は翌月を表示
        if now.day >= 21:
            if now.month == 12:
                year = now.year + 1
                month = 1
            else:
                year = now.year
                month = now.month + 1
        else:
            year = now.year
            month = now.month

    # 前月の21日から当月の20日までの日付範囲を取得
    if month == 1:
        start_date = datetime(year - 1, 12, 21)
    else:
        start_date = datetime(year, month - 1, 21)
    end_date = datetime(year, month, 20)
    
    # 全従業員を取得（雇用形態と従業員IDでソート）
    employees = User.query.order_by(
        case(
            (User.employment_type == 'FT', 0),
            (User.employment_type == 'PT', 1),
        ),
        User.employee_id
    ).all()
    
    # 日付範囲の生成
    dates = []
    current_date = start_date
    while current_date <= end_date:
        dates.append({
            'date': current_date.day,
            'day': ['日', '月', '火', '水', '木', '金', '土'][current_date.weekday()]
        })
        current_date += timedelta(days=1)
    
    # 従業員の勤怠データを取得
    attendance_data = []
    for employee in employees:
        employee_data = {
            'name': employee.name,
            'employment_type': employee.employment_type,
            'attendance': {}
        }
        
        current_date = start_date
        while current_date <= end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            records = AttendanceRecord.query.filter_by(
                user_id=employee.id,
                date=date_str
            ).order_by(AttendanceRecord.time).all()
            
            day_data = {}
            for record in records:
                if record.action_type == '出勤':
                    day_data['arrive'] = record.time.strftime('%-H:%M')
                elif record.action_type == '退勤':
                    day_data['leave'] = record.time.strftime('%-H:%M')
                elif record.action_type == '外出':
                    day_data['out'] = record.time.strftime('%-H:%M')
                elif record.action_type == '戻り':
                    day_data['back'] = record.time.strftime('%-H:%M')
            
            if day_data:
                employee_data['attendance'][str(current_date.day)] = day_data
            
            current_date += timedelta(days=1)
        
        attendance_data.append(employee_data)
    
    return render_template('testapp/attendance_history.html', 
                         year=year, 
                         month=month,
                         dates=dates,
                         attendance_data=attendance_data)

# 勤怠履歴画面のデータ取得
@app.route('/get_attendance_history')
def get_attendance_history():
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    # 21日から翌月20日までの日付範囲を取得
    start_date = datetime(year, month, 21)
    if month == 12:
        end_date = datetime(year + 1, 1, 20)
    else:
        end_date = datetime(year, month + 1, 20)
    
    # 全従業員を取得（雇用形態と従業員IDでソート）
    employees = User.query.order_by(
        case(
            (User.employment_type == 'FT', 0),
            (User.employment_type == 'PT', 1),
        ),
        User.employee_id
    ).all()
    
    result = []
    
    for employee in employees:
        employee_data = {
            'employee_id': employee.employee_id,
            'name': employee.name,
            'employment_type': employee.employment_type,
            'attendance': {}
        }
        
        # 各日の勤怠データを取得
        current_date = start_date
        while current_date <= end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            records = AttendanceRecord.query.filter_by(
                user_id=employee.id,  # user_idを使用
                date=date_str
            ).order_by(AttendanceRecord.time).all()
            
            day_data = {}
            for record in records:
                if record.action_type == '出勤':
                    day_data['arrive'] = record.time.strftime('%-H:%M')
                elif record.action_type == '退勤':
                    day_data['leave'] = record.time.strftime('%-H:%M')
                elif record.action_type == '外出':
                    day_data['out'] = record.time.strftime('%-H:%M')
                elif record.action_type == '戻り':
                    day_data['back'] = record.time.strftime('%-H:%M')
            
            if day_data:
                employee_data['attendance'][str(current_date.day)] = day_data
            
            current_date += timedelta(days=1)
        
        result.append(employee_data)
    
    return jsonify(result)

# 勤怠履歴画面のPDFダウンロード
@app.route('/download_pdf/<int:year>/<int:month>')
def download_pdf(year, month):
    # 前月の21日から当月の20日までの日付範囲を取得
    if month == 1:
        start_date = datetime(year - 1, 12, 21)
    else:
        start_date = datetime(year, month - 1, 21)
    end_date = datetime(year, month, 20)
    
    # 全従業員を取得（雇用形態と従業員IDでソート）
    employees = User.query.order_by(
        case(
            (User.employment_type == 'FT', 0),
            (User.employment_type == 'PT', 1),
        ),
        User.employee_id
    ).all()
    
    # PDFの生成
    buffer = BytesIO()
    # A4横向きのページサイズを設定
    width, height = A4
    doc = SimpleDocTemplate(
        buffer,  # 横向きにするために幅と高さを入れ替え
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    # メタデータの設定
    doc.title = f"{year}年{month}月 勤怠履歴"
    doc.creator = "勤怠管理システム"  # 任意のアプリケーション名など


    elements = []

    # スタイルの設定
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    # CIDフォントを登録
    pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
    # タイトル用スタイルのフォント名変更
    title_style.fontName = 'HeiseiKakuGo-W5'
    title_style.fontSize = 20
    
    #title = Paragraph(f"{year}年{month}月 勤怠履歴（{start_date.strftime('%m/%d')}〜{end_date.strftime('%m/%d')}）", title_style)
    # タイトルの追加
    title = Paragraph(f"{year}年{month}月 勤怠履歴", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))

    # テーブルデータの準備
    # ヘッダー行の作成
    headers = ['従業員名']
    current_date = start_date
    while current_date <= end_date:
        headers.append(f"{current_date.day}\n({['月', '火', '水', '木', '金', '土', '日'][current_date.weekday()]})")
        current_date += timedelta(days=1)
    data = [headers]

    # 各従業員のデータを追加
    for employee in employees:
        row = [employee.name]  
        current_date = start_date
        while current_date <= end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            records = AttendanceRecord.query.filter_by(
                user_id=employee.id,
                date=date_str
            ).order_by(AttendanceRecord.time).all()
            
            if not records:
                row.append('')  # データがない場合は空白
            else:
                if employee.employment_type == 'FT':
                    # FTの場合は出勤時間と退勤時間のみを縦に表示
                    arrive = next((r.time.strftime('%-H:%M') for r in records if r.action_type == '出勤'), '')
                    leave = next((r.time.strftime('%-H:%M') for r in records if r.action_type == '退勤'), '')
                    cell_data = []
                    if arrive:
                        cell_data.append(arrive)
                    if leave:
                        cell_data.append(leave)
                else:
                    # PTの場合は出勤、外出、戻り、退勤を縦に表示
                    arrive = next((r.time.strftime('%-H:%M') for r in records if r.action_type == '出勤'), '')
                    out = next((r.time.strftime('%-H:%M') for r in records if r.action_type == '外出'), '')
                    back = next((r.time.strftime('%-H:%M') for r in records if r.action_type == '戻り'), '')
                    leave = next((r.time.strftime('%-H:%M') for r in records if r.action_type == '退勤'), '')
                    cell_data = []
                    cell_data.append(arrive)
                    cell_data.append(out)  # 外出の記録がない場合は空文字列が追加される
                    cell_data.append(back)  # 戻りの記録がない場合は空文字列が追加される
                    cell_data.append(leave)
                
                row.append('\n'.join(cell_data) if cell_data else '')
            
            current_date += timedelta(days=1)
        data.append(row)

    # テーブルの作成
    # 列幅を調整（従業員名の列を広く、日付の列を適度な幅に）
    col_widths = [55] + [25] * (len(headers) - 1)
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # テーブルのスタイル設定
    table.setStyle(TableStyle([
        # ヘッダー行のスタイル
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'HeiseiKakuGo-W5'),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        
        # データ行のスタイル
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),  # 全ての列を中央揃えに
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),  # 垂直方向も中央揃えに
        ('FONTNAME', (0, 1), (-1, -1), 'HeiseiKakuGo-W5'),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        
        # グリッドの設定
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        
        # セルのパディング
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        
        # 交互の行の背景色
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))

    # 出勤または退勤の記録がないセルを赤字に設定
    for i, employee in enumerate(employees, 1):
        current_date = start_date
        col = 1  # 列インデックス（0から始まる）
        while current_date <= end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            records = AttendanceRecord.query.filter_by(
                user_id=employee.id,
                date=date_str
            ).order_by(AttendanceRecord.time).all()
            
            if records:
                arrive = next((r.time.strftime('%-H:%M') for r in records if r.action_type == '出勤'), '')
                leave = next((r.time.strftime('%-H:%M') for r in records if r.action_type == '退勤'), '')
                
                # 出勤または退勤の記録がない場合、そのセルを赤字に設定
                if (arrive and not leave) or (not arrive and leave):
                    table.setStyle(TableStyle([
                        ('TEXTCOLOR', (col, i), (col, i), colors.red)
                    ]))
            
            col += 1
            current_date += timedelta(days=1)


    elements.append(table)
    
    # PDFの生成
    doc.build(elements)

    # PDFをレスポンスとして返す
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'勤怠履歴_{year}年{month}月.pdf',
        mimetype='application/pdf'
    )

# 月次集計表画面
@app.route('/report')
@app.route('/report/<int:year>/<int:month>')
def report(year=None, month=None):
    # パラメータが指定されていない場合は現在の日付を使用
    if year is None or month is None:
        now = datetime.now()
        # 21日以降は翌月を表示
        if now.day >= 21:
            if now.month == 12:
                year = now.year + 1
                month = 1
            else:
                year = now.year
                month = now.month + 1
        else:
            year = now.year
            month = now.month

    # 前月の21日から当月の20日までの日付範囲を取得
    if month == 1:
        start_date = datetime(year - 1, 12, 21)
    else:
        start_date = datetime(year, month - 1, 21)
    end_date = datetime(year, month, 20)
    
    # 全従業員を取得（雇用形態と従業員IDでソート）
    employees = User.query.order_by(
        case(
            (User.employment_type == 'FT', 0),
            (User.employment_type == 'PT', 1),
        ),
        User.employee_id
    ).all()
    
    # 従業員の勤怠データを取得
    report_data = []
    for employee in employees:
        current_date = start_date
        while current_date <= end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            records = AttendanceRecord.query.filter_by(
                user_id=employee.id,
                date=date_str
            ).order_by(AttendanceRecord.time).all()
            
            # 勤怠情報の取得
            arrive = next((r.time.strftime('%-H:%M') for r in records if r.action_type == '出勤'), '')
            leave = next((r.time.strftime('%-H:%M') for r in records if r.action_type == '退勤'), '')
            
            # 勤怠情報の有無に関わらずデータを追加
            report_data.append({
                'employee_id': employee.employee_id,
                'name': employee.name,
                'year': current_date.year,
                'month': current_date.month,
                'day': current_date.day,
                'arrive': arrive,
                'leave': leave
            })
            
            current_date += timedelta(days=1)
    
    return render_template('testapp/report.html', 
                         year=year, 
                         month=month,
                         report_data=report_data)

@app.route('/prev_month_report/<int:year>/<int:month>')
def prev_month_report(year, month):
    if month == 1:
        return redirect(url_for('report', year=year-1, month=12))
    else:
        return redirect(url_for('report', year=year, month=month-1))

# 翌月の月次集計表画面
@app.route('/next_month_report/<int:year>/<int:month>')
def next_month_report(year, month):
    if month == 12:
        return redirect(url_for('report', year=year+1, month=1))
    else:
        return redirect(url_for('report', year=year, month=month+1))

# 従業員IDとパスワードのチェック
@app.route('/check_employee_id', methods=['POST'])
def check_employee_id():
    data = request.get_json()
    employee_id = data.get('employee_id')
    password = data.get('password')
    
    user = User.query.filter_by(employee_id=employee_id).first()
    
    if not user:
        return jsonify({'exists': False, 'error': '存在しない従業員IDです'})
    
    if not check_password_hash(user.password, password):
        return jsonify({'exists': True, 'error': 'パスワードが異なります'})
    
    return jsonify({'exists': True})

# アカウント削除
@app.route('/delete_account/<int:user_id>', methods=['POST'])
def delete_account(user_id):
    try:
        # ユーザーを取得
        user = User.query.get_or_404(user_id)
        
        # ユーザーを削除
        db.session.delete(user)
        db.session.commit()
        
        # 成功メッセージをフラッシュ
        flash('アカウントを削除しました。', 'success')
        
    except Exception as e:
        # エラーが発生した場合はロールバック
        db.session.rollback()
        # エラーメッセージをフラッシュ
        flash(f'アカウントの削除に失敗しました: {str(e)}', 'error')
    
    # アカウント管理画面にリダイレクト
    return redirect(url_for('management_account'))

# 勤怠登録画面のボタンの状態確認
def check_button_state(user_id, action_type):
    today = datetime.now().date()
    records = AttendanceRecord.query.filter(
        AttendanceRecord.user_id == user_id,
        AttendanceRecord.date == today
    ).order_by(AttendanceRecord.time).all()
    
    # ケース1：記録なし
    if not records:
        return 'confirm'  # 確認が必要
    
    # 最後の記録を取得
    last_record = records[-1]
    
    # ケース2：出勤済み
    if last_record.action_type == '出勤':
        if action_type == '出勤':
            return 'invalid'  # 出勤は無効
        elif action_type in ['外出', '戻り', '退勤']:
            return 'confirm'  # 外出、戻り、退勤は確認が必要
    
    # ケース3：外出済み
    elif last_record.action_type == '外出':
        if action_type in ['出勤', '外出']:
            return 'invalid'  # 出勤と外出は無効
        elif action_type in ['戻り', '退勤']:
            return 'confirm'  # 戻りと退勤は確認が必要
    
    # ケース4：戻り済み
    elif last_record.action_type == '戻り':
        if action_type in ['出勤', '外出', '戻り']:
            return 'invalid'  # 出勤、外出、戻りは無効
        elif action_type == '退勤':
            return 'confirm'  # 退勤は確認が必要
    
    # ケース5：退勤済み
    elif last_record.action_type == '退勤':
        return 'invalid'  # 全てのボタンは無効
    
    return 'invalid'

# 勤怠登録画面の記録
@app.route('/record_attendance', methods=['POST'])
def record_attendance():
    if 'user_id' not in session:
        return redirect(url_for('employee_login'))
    
    user_id = session['user_id']
    action_type = request.form.get('action_type')
    
    # ボタンの状態を確認
    state = check_button_state(user_id, action_type)
    
    if state == 'confirm':
        # モーダルで確認するため、エラーメッセージを返す
        return jsonify({'status': 'confirm', 'message': f'{action_type}を記録しますか？'})
    
    else:
        # エラーメッセージを返す
        return jsonify({'status': 'error', 'message': 'この操作は現在無効です。'})

# 勤怠登録画面の確認
@app.route('/confirm_attendance', methods=['POST'])
def confirm_attendance():
    if 'user_id' not in session:
        return redirect(url_for('employee_login'))
    
    user_id = session['user_id']
    action_type = request.form.get('action_type')
    
    # 確認後の記録
    JST = timezone(timedelta(hours=9), 'JST')
    current_time = datetime.now(JST)
    record = AttendanceRecord(
        user_id=user_id,
        action_type=action_type,
        date=current_time.date(),
        time=current_time.time(),
        day_of_week=current_time.weekday()
    )
    db.session.add(record)
    db.session.commit()
    
    return jsonify({'status': 'success', 'message': f'{action_type}を記録しました。'})

# 勤怠登録画面の記録
@app.route('/get_todays_records')
def get_todays_records():
    if 'user_id' not in session:
        return jsonify([])
    
    user_id = session['user_id']
    today = datetime.now().date()
    
    records = AttendanceRecord.query.filter(
        AttendanceRecord.user_id == user_id,
        AttendanceRecord.date == today
    ).order_by(AttendanceRecord.time).all()
    
    return jsonify([{
        'action_type': record.action_type,
        'time': record.time.strftime('%H:%M')
    } for record in records])

# 管理者IDとパスワードのチェック
@app.route('/check_adm_id', methods=['POST'])
def check_adm_id():
    data = request.get_json()
    adm_id = data.get('adm_id')
    adm_password = data.get('adm_password')
    
    if adm_id != ADMIN_ID:
        return jsonify({'exists': False, 'error': '管理者IDが異なります。'})
    
    if adm_password != ADMIN_PASSWORD:
        return jsonify({'exists': True, 'error': 'パスワードが異なります。'})
    
    return jsonify({'exists': True})

# 正社員用の勤怠登録画面のボタンの状態確認
def check_button_state_FT(user_id, action_type):
    today = datetime.now().date()
    records = AttendanceRecord.query.filter(
        AttendanceRecord.user_id == user_id,
        AttendanceRecord.date == today
    ).order_by(AttendanceRecord.time).all()
    
    # ケース1：記録なし
    if not records:
        return 'confirm'  # 確認が必要
    
    # 最後の記録を取得
    last_record = records[-1]
    
    # ケース2：出勤済み
    if last_record.action_type == '出勤':
        if action_type == '出勤':
            return 'invalid'  # 出勤は有効
        elif action_type == '退勤':
            return 'confirm'  # 退勤は確認が必要
    
    # ケース3：退勤済み
    elif last_record.action_type == '退勤':
        return 'invalid'  # 全てのボタンは無効
    
    return 'invalid'

# 正社員用の勤怠登録画面の記録
@app.route('/record_attendance_FT', methods=['POST'])
def record_attendance_FT():
    if 'user_id' not in session:
        return redirect(url_for('employee_login'))
    
    user_id = session['user_id']
    action_type = request.form.get('action_type')
    
    # ボタンの状態を確認
    state = check_button_state_FT(user_id, action_type)
    
    if state == 'confirm':
        # モーダルで確認するため、エラーメッセージを返す
        return jsonify({'status': 'confirm', 'message': f'{action_type}を記録しますか？'})
    elif state == 'invalid':
        return jsonify({'status': 'invalid'})
    else:
        return jsonify({'status': 'error', 'message': '無効な操作です。'})

# パート用の勤怠登録画面の記録
@app.route('/record_attendance_PT', methods=['POST'])
def record_attendance_PT():
    if 'user_id' not in session:
        return redirect(url_for('employee_login'))
    
    user_id = session['user_id']
    action_type = request.form.get('action_type')
    
    # ボタンの状態を確認
    state = check_button_state(user_id, action_type)
    
    if state == 'confirm':
        return jsonify({'status': 'confirm'})
    elif state == 'invalid':
        return jsonify({'status': 'invalid'})
    else:
        return jsonify({'status': 'error', 'message': '無効な操作です。'})

# 勤怠履歴画面の前月の画面
@app.route('/prev_month/<int:year>/<int:month>')
def prev_month(year, month):
    if month == 1:
        return redirect(url_for('attendance_history', year=year-1, month=12))
    else:
        return redirect(url_for('attendance_history', year=year, month=month-1))

# 勤怠履歴画面の翌月の画面
@app.route('/next_month/<int:year>/<int:month>')
def next_month(year, month):
    if month == 12:
        return redirect(url_for('attendance_history', year=year+1, month=1))
    else:
        return redirect(url_for('attendance_history', year=year, month=month+1))

# 月次集計表画面のCSVダウンロード
@app.route('/download_csv/<int:year>/<int:month>')
def download_csv(year, month):
    # 前月の21日から当月の20日までの日付範囲を取得
    if month == 1:
        start_date = datetime(year - 1, 12, 21)
    else:
        start_date = datetime(year, month - 1, 21)
    end_date = datetime(year, month, 20)
    
    # 全従業員を取得（雇用形態と従業員IDでソート）
    employees = User.query.order_by(
        case(
            (User.employment_type == 'FT', 0),
            (User.employment_type == 'PT', 1),
        ),
        User.employee_id
    ).all()
    
    # Excelファイルの生成
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月 月次集計表"
    
    # スタイルの設定
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ヘッダー行
    headers = ['従業員番号', '従業員名', '年', '月', '日', '出勤時間', '退勤時間']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # データ行
    row_num = 2
    for employee in employees:
        current_date = start_date
        first_row = True
        
        while current_date <= end_date:
            date_str = current_date.strftime('%Y-%m-%d')
            records = AttendanceRecord.query.filter_by(
                user_id=employee.id,
                date=date_str
            ).order_by(AttendanceRecord.time).all()
            
            # 勤怠情報の取得
            arrive = next((r.time.strftime('%-H:%M') for r in records if r.action_type == '出勤'), '')
            leave = next((r.time.strftime('%-H:%M') for r in records if r.action_type == '退勤'), '')
            
            # Excel行の追加
            if first_row:
                ws.cell(row=row_num, column=1, value=employee.employee_id)
                ws.cell(row=row_num, column=2, value=employee.name)
                first_row = False
            else:
                ws.cell(row=row_num, column=1, value='')
                ws.cell(row=row_num, column=2, value='')
            
            ws.cell(row=row_num, column=3, value=current_date.year)
            ws.cell(row=row_num, column=4, value=current_date.month)
            ws.cell(row=row_num, column=5, value=current_date.day)
            ws.cell(row=row_num, column=6, value=arrive)
            ws.cell(row=row_num, column=7, value=leave)
            
            # セルのスタイル設定
            for col in range(1, 8):
                cell = ws.cell(row=row_num, column=col)
                cell.alignment = center_alignment
                cell.border = thin_border
            
            row_num += 1
            current_date += timedelta(days=1)
        
        # 従業員ごとに1行の空白を追加
        row_num += 1
    
    # 列幅の調整
    ws.column_dimensions['A'].width = 15  # 従業員番号
    ws.column_dimensions['B'].width = 20  # 従業員名
    ws.column_dimensions['C'].width = 10  # 年度
    ws.column_dimensions['D'].width = 10  # 月
    ws.column_dimensions['E'].width = 10  # 日
    ws.column_dimensions['F'].width = 15  # 出勤時間
    ws.column_dimensions['G'].width = 15  # 退勤時間
    
    # Excelファイルをレスポンスとして返す
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'月次集計表_{year}年{month}月.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    # 開発環境の場合はデバッグモードを有効化
    app.run(debug=True)
   